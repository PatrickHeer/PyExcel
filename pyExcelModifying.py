package EqOrder

from  openpyxl import load_workbook
import openpyxl
##filename = 'C:\Users\patrickh\73INV.xlsx'

##wb_obj = openpyxl.load_workbook(path)

#sheet_obj = wb_obj.active

##cell_obj = sheet_obj.cell(row = 1, column = 1)

##cell_obj = sheet_obj.cell(row = 1, column = 8)

path = r'C:\Users\patrickh\Usage321521.xlsx'

wb = load_workbook(path)

ws = wb['Usage321521']
all_rows = list(ws.rows)
all_columns = list(ws.columns)

d = {}


for row in all_rows[0:204]:
    model = row[0].value
    issued = row[7].value
    description = row[3].value
    string = "Daikin"
    for cell in row:
        try:
            if "Daikin" in cell.value or "DAIKIN" in cell.value or "daikin" in cell.value:                  
                print(f"{model}")
                print(f"{issued}")
                d[f"{model}"] = f"{issued}"
        except (AttributeError, TypeError):
            continue

###################################################
path2 = r'C:\Users\patrickh\DaikinOrderTemplate.xlsx'

wb2 = load_workbook(path2)

ws2 = wb2['DaikinOrderTemplate']

mr = ws.max_row
mc = ws.max_column




for i in range(7,mr +1):
    for j in range(1,mc +1):

        try: 
            c = ws.cell(row = i, column = 1)
            f = ws.cell(row = i, column = 9)
            c2 = ws2.cell(row = i, column = 1)
           
                
            if c2.value in d:
                for row in all_rows[7:204]:    
                    ws2.cell(row = i, column = 11).value = d.get(c2.value)
        except (AttributeError, TypeError):
            continue


wb2.save(str('DaikinOrderTemplate'))        
